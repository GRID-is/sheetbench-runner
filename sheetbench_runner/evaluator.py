"""
Workbook evaluation for SpreadsheetBench.

Ported from SpreadsheetBench/evaluation/evaluation_verified.py with the same
comparison logic to ensure compatible results.
"""

import datetime
import re
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils.cell import SHEETRANGE_RE

from .entities import EvaluationResult, Task


def _datetime_to_excel_serial(dt: datetime.datetime) -> float:
    """Convert a datetime to Excel serial date number."""
    excel_start_date = datetime.datetime(1899, 12, 30)
    delta = dt - excel_start_date
    return delta.days + delta.seconds / 86400.0


def _transform_value(value: Any) -> Any:
    """
    Transform a cell value for comparison.

    Applies the same transformations as SpreadsheetBench evaluation:
    - Numbers: round to 2 decimal places
    - datetime.time: convert to string without milliseconds
    - datetime.datetime: convert to Excel serial number (rounded to 0 decimals)
    - Strings that look like numbers: parse and round to 2 decimals
    """
    if isinstance(value, (int, float)):
        return round(float(value), 2)
    elif isinstance(value, datetime.time):
        return str(value)[:-3]  # Remove milliseconds
    elif isinstance(value, datetime.datetime):
        return round(_datetime_to_excel_serial(value), 0)
    elif isinstance(value, str):
        try:
            return round(float(value), 2)
        except ValueError:
            pass
    return value


def _compare_cell_values(v1: Any, v2: Any) -> bool:
    """
    Compare two cell values after transformation.

    Empty string and None are treated as equivalent.
    """
    v1 = _transform_value(v1)
    v2 = _transform_value(v2)

    # Empty/None equivalence
    if (v1 == "" and v2 is None) or (v1 is None and v2 == ""):
        return True
    if (v1 == "" and v2 == "") or (v1 is None and v2 is None):
        return True

    # Type mismatch
    if type(v1) is not type(v2):
        return False

    return bool(v1 == v2)


def _col_name_to_num(name: str) -> int:
    """Convert an Excel column name (e.g., 'AB') to a number (1-indexed)."""
    num = 0
    for c in name:
        num = num * 26 + (ord(c) - ord("A") + 1)
    return num


def _col_num_to_name(n: int) -> str:
    """Convert a column number (1-indexed) to an Excel column name."""
    name = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        name = chr(65 + remainder) + name
    return name


def _parse_cell_range(range_str: str) -> tuple[tuple[int, int], tuple[int, int]]:
    """
    Parse a range string like 'A1:AB12'.

    Returns ((start_col, start_row), (end_col, end_row)) with 1-indexed values.
    """
    start_cell, end_cell = range_str.split(":")

    start_col, start_row = "", ""
    for char in start_cell:
        if char.isdigit():
            start_row += char
        else:
            start_col += char

    end_col, end_row = "", ""
    for char in end_cell:
        if char.isdigit():
            end_row += char
        else:
            end_col += char

    return (
        (_col_name_to_num(start_col), int(start_row)),
        (_col_name_to_num(end_col), int(end_row)),
    )


def _generate_cell_names(range_str: str) -> list[str]:
    """Generate a list of all cell names in the specified range."""
    if ":" not in range_str:
        return [range_str]

    (start_col, start_row), (end_col, end_row) = _parse_cell_range(range_str)
    columns = [_col_num_to_name(i) for i in range(start_col, end_col + 1)]
    return [f"{col}{row}" for col in columns for row in range(start_row, end_row + 1)]


def _repair_unbalanced_quotes(s: str) -> str:
    """
    Attempt to repair common quote issues in Excel range strings.

    Handles cases like:
    - "Sheet Name'!A1" -> "'Sheet Name'!A1" (missing opening quote)
    - "'9045!A1:F9'" -> "'9045'!A1:F9" (quote wrapping entire reference)
    """
    # Count unescaped quotes (escaped = doubled '')
    quote_count = s.count("'") - s.count("''") * 2
    if quote_count % 2 == 0:
        return s  # Balanced, no repair needed

    # Pattern 1: Missing opening quote before sheet name
    # Must match at segment boundary (start of string or after comma)
    # "SOME SHEET'!A1" -> "'SOME SHEET'!A1"
    # Captures: (boundary)(sheet_name)('!)
    s = re.sub(r"(^|,\s*)([A-Za-z0-9][A-Za-z0-9 ]*)'!", r"\1'\2'!", s)

    return s


def _split_on_comma_outside_quotes(s: str) -> list[str]:
    """
    Split a string on commas, but only when they are outside quoted sections.

    Handles both single and double quotes. Quotes inside the string are expected
    to be balanced (e.g., "'Sheet, Name'!A1" has balanced single quotes).

    Examples:
        "A1,B2" -> ["A1", "B2"]
        "'a, b'!A1,Sheet2!B2" -> ["'a, b'!A1", "Sheet2!B2"]
    """
    parts: list[str] = []
    current: list[str] = []
    in_single_quote = False
    in_double_quote = False

    for char in s:
        if char == "'" and not in_double_quote:
            in_single_quote = not in_single_quote
            current.append(char)
        elif char == '"' and not in_single_quote:
            in_double_quote = not in_double_quote
            current.append(char)
        elif char == "," and not in_single_quote and not in_double_quote:
            parts.append("".join(current))
            current = []
        else:
            current.append(char)

    # Don't forget the last part
    parts.append("".join(current))
    return parts


def _parse_sheet_cell_ranges(
    answer_position: str,
    answer_sheet: str | None = None,
    fallback_sheet_name: str | None = None,
) -> list[tuple[str, str]]:
    """
    Parse answer_position into a list of (sheet_name, cell_range) tuples.

    Args:
        answer_position: The answer position string, potentially with multiple
            comma-separated ranges and optional sheet names (e.g., "Sheet1!A1:B10").
            Sheet names may contain commas if quoted (e.g., "'Sales, Q1'!A1").
            Handles malformed quotes (missing opening quote) via repair.
        answer_sheet: Optional answer_sheet field from the task.
        fallback_sheet_name: Sheet name to use if none specified.

    Returns:
        List of (sheet_name, cell_range) tuples.
    """
    # Repair common quote issues before parsing
    answer_position = _repair_unbalanced_quotes(answer_position)

    # Split on commas, but respect quoted sheet names
    sheet_cell_ranges = _split_on_comma_outside_quotes(answer_position)
    result: list[tuple[str, str]] = []

    for part in sheet_cell_ranges:
        part = part.strip()

        if "!" in part:
            # Try parsing with openpyxl's regex first
            match = SHEETRANGE_RE.match(part)
            if match:
                sheet_name = match.group("quoted") or match.group("notquoted")
                # Unescape doubled quotes
                sheet_name = sheet_name.replace("''", "'")
                cell_range = match.group("cells")
            else:
                # Fallback: handle edge cases like external workbook refs
                # "[workbook.xlsx]'sheet'!A1" or malformed input
                sheet_name, cell_range = part.rsplit("!", 1)
                # Handle external ref: "[book.xlsx]'sheet'" -> "[book.xlsx]sheet"
                if sheet_name.startswith("[") and "'" in sheet_name:
                    # Strip quotes from sheet name portion after the ]
                    bracket_end = sheet_name.find("]")
                    if bracket_end != -1:
                        prefix = sheet_name[: bracket_end + 1]
                        suffix = sheet_name[bracket_end + 1 :].strip("'\"")
                        sheet_name = prefix + suffix
                else:
                    sheet_name = sheet_name.strip("'\"")
        elif answer_sheet:
            sheet_name = answer_sheet.strip().strip("'\"")
            cell_range = part
        elif fallback_sheet_name:
            sheet_name = fallback_sheet_name
            cell_range = part
        else:
            raise ValueError("No sheet name available")

        cell_range = cell_range.strip("'\"")
        result.append((sheet_name, cell_range))

    return result


def _compare_cells(
    wb_golden: openpyxl.Workbook,
    wb_output: openpyxl.Workbook,
    sheet_name: str,
    cell_range: str,
) -> tuple[bool, str]:
    """
    Compare cells in a specific range between two workbooks.

    Returns (passed, message) where message explains any failure.
    """
    if sheet_name not in wb_output.sheetnames:
        return False, f"Worksheet '{sheet_name}' not found in output"
    if sheet_name not in wb_golden.sheetnames:
        return False, f"Worksheet '{sheet_name}' not found in golden file"

    ws_golden = wb_golden[sheet_name]
    ws_output = wb_output[sheet_name]

    cell_names = _generate_cell_names(cell_range)

    for cell_name in cell_names:
        cell_golden = ws_golden[cell_name]
        cell_output = ws_output[cell_name]

        if not _compare_cell_values(cell_golden.value, cell_output.value):
            return (
                False,
                f"Value mismatch at {sheet_name}!{cell_name}: "
                f"expected {cell_golden.value!r}, got {cell_output.value!r}",
            )

    return True, ""


class Evaluator:
    """Evaluates task outputs against golden files."""

    def __init__(self, dataset_path: Path):
        """
        Initialize the evaluator.

        Args:
            dataset_path: Path to the SpreadsheetBench dataset directory
        """
        self.dataset_path = dataset_path

    def evaluate(self, task: Task, output_path: Path) -> EvaluationResult:
        """
        Evaluate a task output against its golden file.

        Args:
            task: The task that was executed
            output_path: Path to the output spreadsheet file

        Returns:
            EvaluationResult with passed status and message
        """
        if not output_path.exists():
            return EvaluationResult(passed=False, message="Output file not found")

        golden_path = self._get_golden_path(task)
        if not golden_path.exists():
            return EvaluationResult(passed=False, message=f"Golden file not found: {golden_path}")

        try:
            return self._compare_workbooks(task, golden_path, output_path)
        except Exception as e:
            return EvaluationResult(passed=False, message=f"Evaluation error: {e}")

    def _get_golden_path(self, task: Task) -> Path:
        """Get the path to the golden file for a task."""
        return self.dataset_path / "spreadsheet" / task.id / f"1_{task.id}_golden.xlsx"

    def _compare_workbooks(
        self, task: Task, golden_path: Path, output_path: Path
    ) -> EvaluationResult:
        """Compare the output workbook against the golden file."""
        wb_golden = openpyxl.load_workbook(filename=golden_path, data_only=True)
        wb_output = openpyxl.load_workbook(filename=output_path, data_only=True)

        all_passed = True
        messages: list[str] = []

        parsed_ranges = _parse_sheet_cell_ranges(
            task.answer_position,
            task.answer_sheet,
            wb_golden.sheetnames[0],
        )

        for sheet_name, cell_range in parsed_ranges:
            passed, message = _compare_cells(wb_golden, wb_output, sheet_name, cell_range)
            if not passed:
                all_passed = False
                messages.append(message)

        wb_golden.close()
        wb_output.close()

        return EvaluationResult(
            passed=all_passed,
            message="; ".join(m for m in messages if m),
        )
