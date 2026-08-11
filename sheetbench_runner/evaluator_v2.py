"""
V2 evaluation for SpreadsheetBench v2 datasets (Debugging, Financial_Model, Template).

Ported from SpreadsheetBench-2/evaluation/evaluation.py and kept structurally
parallel to it so the two can be diffed. Semantics-bearing functions are
vendored; generic helpers (_transform_value, _generate_cell_names) are reused
from evaluator.py. This module must not be imported at module level by
evaluator.py (evaluator.py uses a function-local import for dispatch) so the
helper imports below stay acyclic.
"""

from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Font
from openpyxl.styles.colors import Color
from openpyxl.worksheet.worksheet import Worksheet

from .evaluator import _generate_cell_names, _transform_value

_DISPLAY_EQUIVALENT_ERRORS = {"#DIV/0!", "#N/A"}
# Finance "not meaningful" placeholders: golden #DIV/0! vs output "N/A" (via
# IFERROR) must match.
_NOT_MEANINGFUL = _DISPLAY_EQUIVALENT_ERRORS | {
    "N/A",
    "NA",
    "N.A.",
    "N/M",
    "NM",
    "N.M.",
    "NOT MEANINGFUL",
    "NOT APPLICABLE",
    "NOT AVAILABLE",
    "—",
    "–",
    "-",
    "--",
    "---",
}


def _is_not_meaningful(v: Any) -> bool:
    if isinstance(v, str):
        return v.strip().upper() in _NOT_MEANINGFUL
    return False


def compare_cell_value(v1: Any, v2: Any, tolerance: float = 0.01) -> bool:
    """Tolerant value comparison (upstream compare_cell_value)."""
    # ArrayFormula objects compare by formula text
    if hasattr(v1, "text") and hasattr(v2, "text"):
        return bool(v1.text == v2.text)

    if _is_not_meaningful(v1) and _is_not_meaningful(v2):
        return True

    # Numeric vs numeric: raw values with tolerance only (no rounding, to
    # avoid boundary artifacts like -0.105 vs -0.10500000000000001)
    if isinstance(v1, (int, float)) and isinstance(v2, (int, float)):
        return _numbers_match(v1, v2, tolerance)

    v1 = _transform_value(v1)
    v2 = _transform_value(v2)
    if (v1 == "" and v2 is None) or (v1 is None and v2 == ""):
        return True
    if (v1 == "" and v2 == "") or (v1 is None and v2 is None):
        return True
    # None is equivalent to 0 (empty cells read as None; rounding can produce
    # 0 from near-zero values)
    if (v1 is None and isinstance(v2, (int, float)) and v2 == 0) or (
        v2 is None and isinstance(v1, (int, float)) and v1 == 0
    ):
        return True
    if type(v1) is not type(v2):
        return False
    if v1 == v2:
        return True
    # Formula strings: Excel function names are case-insensitive
    if isinstance(v1, str) and isinstance(v2, str) and v1.startswith("=") and v2.startswith("="):
        return v1.replace("$", "").upper() == v2.replace("$", "").upper()
    # Numeric-looking strings arrive here as parsed floats
    if isinstance(v1, (int, float)) and isinstance(v2, (int, float)):
        return _numbers_match(v1, v2, tolerance)
    return False


def _numbers_match(v1: float, v2: float, tolerance: float) -> bool:
    if v1 == v2:
        return True
    if v1 == 0 or v2 == 0:
        return abs(v1 - v2) <= tolerance
    return abs(v1 - v2) / max(abs(v1), abs(v2)) <= tolerance


_EXCEL_ERRORS = {"#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#NULL!", "#N/A", "#NUM!"}


def _has_excel_error(value: Any) -> bool:
    """True if the value is an Excel error string."""
    return isinstance(value, str) and value in _EXCEL_ERRORS


def _find_sheet(wb: openpyxl.Workbook, name: str) -> Worksheet | None:
    """Find a worksheet with whitespace-tolerant, case-insensitive matching."""
    if name in wb.sheetnames:
        return wb[name]
    name_stripped = name.strip().lower()
    for sn in wb.sheetnames:
        if sn.strip().lower() == name_stripped:
            return wb[sn]
    return None


def compare_cell_formula(f1: Any, f2: Any) -> bool:
    """Compare formula-level cell values (upstream compare_cell_formula)."""
    # ArrayFormula objects (CSE array formulas) compare by formula text
    if hasattr(f1, "text") and hasattr(f2, "text"):
        return bool(f1.text == f2.text)

    # Both formulas: normalize $ markers, case, and the legacy =+ prefix
    if isinstance(f1, str) and isinstance(f2, str) and f1.startswith("=") and f2.startswith("="):

        def _normalize(f: str) -> str:
            f = f.replace("$", "").upper()
            if f.startswith("=+"):
                f = "=" + f[2:]
            return f

        return _normalize(f1) == _normalize(f2)

    empty = (None, "")
    if f1 in empty and f2 in empty:
        return True

    return compare_cell_value(f1, f2)


# Standard Excel theme color map (Office default theme)
_THEME_COLORS = [
    "FFFFFF",  # 0: lt1 (white/light background)
    "000000",  # 1: dk1 (black/dark text)
    "E7E6E6",  # 2: lt2 (light gray)
    "44546A",  # 3: dk2 (dark blue-gray)
    "4472C4",  # 4: accent1
    "ED7D31",  # 5: accent2
    "A5A5A5",  # 6: accent3
    "FFC000",  # 7: accent4
    "5B9BD5",  # 8: accent5
    "70AD47",  # 9: accent6
]


def _get_color_rgb(color: Color | None) -> str:
    """Extract RGB from a color object, resolving theme colors to RGB."""
    if not color:
        return "00000000"
    # openpyxl descriptors return validator objects when the attribute isn't
    # set, so check color.type rather than testing color.theme/color.rgb
    if color.type == "rgb" and isinstance(color.rgb, str):
        return color.rgb
    if color.type == "theme":
        theme_idx = color.theme
        if 0 <= theme_idx < len(_THEME_COLORS):
            base = _THEME_COLORS[theme_idx]
            tint = float(color.tint) if color.tint else 0.0
            if tint != 0:
                r, g, b = int(base[0:2], 16), int(base[2:4], 16), int(base[4:6], 16)
                if tint > 0:
                    r = int(r + (255 - r) * tint)
                    g = int(g + (255 - g) * tint)
                    b = int(b + (255 - b) * tint)
                else:
                    r = int(r * (1 + tint))
                    g = int(g * (1 + tint))
                    b = int(b * (1 + tint))
                r, g, b = min(255, max(0, r)), min(255, max(0, g)), min(255, max(0, b))
                base = f"{r:02X}{g:02X}{b:02X}"
            return "FF" + base
    return "00000000"


def compare_font_color(font_golden: Font, font_output: Font) -> bool:
    """Compare font colors on RGB only, ignoring the alpha channel."""
    rgb1 = _get_color_rgb(font_golden.color)
    rgb2 = _get_color_rgb(font_output.color)
    return rgb1[-6:] == rgb2[-6:]


class _LazyFormulaWorkbooks:
    """
    data_only=False copies of the input/golden/output workbooks, loaded on
    first use. The formula level is only needed when a cell shows an Excel
    error value, so most tasks never pay the extra load.
    """

    def __init__(self, input_path: Path, golden_path: Path, output_path: Path) -> None:
        self._paths: dict[str, Path] = {
            "input": input_path,
            "golden": golden_path,
            "output": output_path,
        }
        self._books: dict[str, openpyxl.Workbook] = {}

    def sheet(self, which: str, sheet_name: str) -> Worksheet | None:
        if which not in self._books:
            self._books[which] = openpyxl.load_workbook(
                filename=self._paths[which], data_only=False
            )
        return _find_sheet(self._books[which], sheet_name)

    def close(self) -> None:
        for wb in self._books.values():
            wb.close()


def _compare_cells(cell1: Any, cell2: Any, with_font_color: bool, with_formula: bool) -> bool:
    if with_formula:
        return compare_cell_formula(cell1.value, cell2.value)
    if with_font_color:
        return compare_cell_value(cell1.value, cell2.value) and compare_font_color(
            cell1.font, cell2.font
        )
    return compare_cell_value(cell1.value, cell2.value)


def classify_cells_by_modification(
    wb_input: openpyxl.Workbook,
    wb_golden: openpyxl.Workbook,
    sheet_name: str,
    cell_range: str,
    with_font_color: bool,
    with_formula: bool,
    formula_books: _LazyFormulaWorkbooks | None,
) -> tuple[list[str], list[str]]:
    """
    Split the range into regression cells (input == golden, must stay
    untouched) and modification cells (the task's actual work).
    """
    ws_input = _find_sheet(wb_input, sheet_name)
    ws_golden = _find_sheet(wb_golden, sheet_name)
    if ws_input is None or ws_golden is None:
        return [], []

    regression: list[str] = []
    modification: list[str] = []
    for cell_name in _generate_cell_names(cell_range):
        cell_in, cell_gold = ws_input[cell_name], ws_golden[cell_name]
        if (
            not with_formula
            and formula_books is not None
            and (_has_excel_error(cell_in.value) or _has_excel_error(cell_gold.value))
        ):
            ws_in_f = formula_books.sheet("input", sheet_name)
            ws_gold_f = formula_books.sheet("golden", sheet_name)
            assert ws_in_f is not None and ws_gold_f is not None
            is_same = compare_cell_formula(ws_in_f[cell_name].value, ws_gold_f[cell_name].value)
        else:
            is_same = _compare_cells(cell_in, cell_gold, with_font_color, with_formula)
        (regression if is_same else modification).append(cell_name)

    return regression, modification


def compare_classified_cells(
    wb_golden: openpyxl.Workbook,
    wb_output: openpyxl.Workbook,
    sheet_name: str,
    regression_cells: list[str],
    modification_cells: list[str],
    with_font_color: bool,
    with_formula: bool,
    formula_books: _LazyFormulaWorkbooks | None,
) -> tuple[int, int, int, int, list[str]]:
    """
    Compare output vs golden for both cell groups.
    Returns (reg_correct, reg_total, mod_correct, mod_total, error_messages).
    A sheet missing from the output scores all its cells wrong.
    """
    if _find_sheet(wb_output, sheet_name) is None:
        return (
            0,
            len(regression_cells),
            0,
            len(modification_cells),
            [f"{sheet_name} worksheet not found"],
        )

    ws_golden = _find_sheet(wb_golden, sheet_name)
    ws_output = _find_sheet(wb_output, sheet_name)

    def count_correct(cells: list[str], label: str) -> tuple[int, list[str]]:
        correct = 0
        errors: list[str] = []
        for name in cells:
            # Narrowed here (not above) so an empty `cells` list never
            # dereferences a None sheet, matching upstream's lazy access.
            assert ws_golden is not None and ws_output is not None
            cell_gold, cell_out = ws_golden[name], ws_output[name]
            if (
                not with_formula
                and formula_books is not None
                and (_has_excel_error(cell_gold.value) or _has_excel_error(cell_out.value))
            ):
                ws_gold_f = formula_books.sheet("golden", sheet_name)
                ws_out_f = formula_books.sheet("output", sheet_name)
                assert ws_gold_f is not None and ws_out_f is not None
                matched = compare_cell_formula(ws_gold_f[name].value, ws_out_f[name].value)
            else:
                matched = _compare_cells(cell_gold, cell_out, with_font_color, with_formula)
            if matched:
                correct += 1
            else:
                gold_val, out_val = cell_gold.value, cell_out.value
                if hasattr(gold_val, "text"):
                    gold_val = gold_val.text
                if hasattr(out_val, "text"):
                    out_val = out_val.text
                errors.append(
                    f"{label} error at {sheet_name}!{name}: answer={gold_val}, output={out_val}"
                )
        return correct, errors

    reg_correct, reg_errors = count_correct(regression_cells, "Regression")
    mod_correct, mod_errors = count_correct(modification_cells, "Modification")

    return (
        reg_correct,
        len(regression_cells),
        mod_correct,
        len(modification_cells),
        reg_errors + mod_errors,
    )
