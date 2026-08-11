"""Tests for v2 (SpreadsheetBench 2) evaluation semantics."""

import openpyxl
import pytest
from openpyxl.styles import Font
from openpyxl.styles.colors import Color

from sheetbench_runner.entities import Task, TaskResult
from sheetbench_runner.evaluator import Evaluator
from sheetbench_runner.evaluator_v2 import (
    _find_sheet,
    _has_excel_error,
    _LazyFormulaWorkbooks,
    classify_cells_by_modification,
    compare_cell_formula,
    compare_cell_value,
    compare_classified_cells,
    compare_font_color,
    compare_workbooks,
)


def build_workbook(path, sheet_name="Model", cells=None, fonts=None):
    """Write an xlsx with the given {cell: value} and optional {cell: Font}."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    for cell, value in (cells or {}).items():
        ws[cell] = value
    for cell, font in (fonts or {}).items():
        ws[cell].font = font
    wb.save(path)
    return path


class TestCompareCellValue:
    """Port of upstream compare_cell_value semantics."""

    # -- numeric tolerance (1% relative, 0.01 absolute at zero) --

    def test_equal_numbers(self):
        assert compare_cell_value(5, 5) is True

    def test_within_relative_tolerance(self):
        assert compare_cell_value(100.0, 100.9) is True  # 0.9% off

    def test_outside_relative_tolerance(self):
        assert compare_cell_value(100.0, 101.1) is False  # 1.1% off

    def test_zero_uses_absolute_tolerance(self):
        assert compare_cell_value(0, 0.009) is True
        assert compare_cell_value(0, 0.011) is False

    def test_no_rounding_boundary_artifact(self):
        # Raw comparison, not round-to-2: these differ by ~1e-17
        assert compare_cell_value(-0.105, -0.10500000000000001) is True

    # -- "not meaningful" equivalence --

    @pytest.mark.parametrize(
        "golden,output",
        [
            ("#DIV/0!", "N/A"),
            ("#N/A", "NM"),
            ("#DIV/0!", "n.m."),
            ("#N/A", "—"),
            ("-", "--"),
            ("Not Meaningful", "#DIV/0!"),
        ],
    )
    def test_not_meaningful_pairs_match(self, golden, output):
        assert compare_cell_value(golden, output) is True

    def test_not_meaningful_vs_number_fails(self):
        assert compare_cell_value("#DIV/0!", 5.0) is False

    def test_plain_string_not_in_equivalence_class(self):
        assert compare_cell_value("N/A", "hello") is False

    # -- empty / None / zero equivalences --

    def test_none_equals_empty_string(self):
        assert compare_cell_value(None, "") is True
        assert compare_cell_value("", None) is True

    def test_none_equals_zero(self):
        assert compare_cell_value(None, 0) is True
        assert compare_cell_value(0.0, None) is True

    def test_none_vs_nonzero_fails(self):
        assert compare_cell_value(None, 5) is False

    # -- strings --

    def test_numeric_strings_compared_with_tolerance(self):
        assert compare_cell_value("100.0", "100.5") is True
        assert compare_cell_value("100.0", "102.0") is False

    def test_numeric_string_vs_number(self):
        assert compare_cell_value("3.14159", 3.14) is True

    def test_formula_strings_case_and_dollar_insensitive(self):
        assert compare_cell_value("=sum($A$1:B2)", "=SUM(A1:B2)") is True
        assert compare_cell_value("=SUM(A1:B2)", "=SUM(A1:B3)") is False

    def test_type_mismatch_fails(self):
        assert compare_cell_value("hello", 5.0) is False


class TestCompareCellFormula:
    def test_identical_formulas(self):
        assert compare_cell_formula("=SUM(A1:B2)", "=SUM(A1:B2)") is True

    def test_case_dollar_and_plus_normalized(self):
        assert compare_cell_formula("=+sum($A$1:B2)", "=SUM(A1:B2)") is True

    def test_different_formulas(self):
        assert compare_cell_formula("=SUM(A1:B2)", "=SUM(A1:B3)") is False

    def test_none_and_empty_equivalent(self):
        assert compare_cell_formula(None, "") is True

    def test_non_formula_values_fall_back_to_value_compare(self):
        assert compare_cell_formula(100.0, 100.9) is True
        assert compare_cell_formula(100.0, 102.0) is False


class TestHasExcelError:
    @pytest.mark.parametrize(
        "value", ["#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#NULL!", "#N/A", "#NUM!"]
    )
    def test_error_strings(self, value):
        assert _has_excel_error(value) is True

    def test_non_errors(self):
        assert _has_excel_error("hello") is False
        assert _has_excel_error(5.0) is False
        assert _has_excel_error(None) is False


class TestFindSheet:
    def test_exact_and_fuzzy_match(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "My Sheet "
        assert _find_sheet(wb, "My Sheet ") is ws
        assert _find_sheet(wb, "my sheet") is ws
        assert _find_sheet(wb, "Other") is None


class TestCompareFontColor:
    def test_same_rgb_matches(self):
        assert compare_font_color(Font(color="FFFF0000"), Font(color="FFFF0000")) is True

    def test_alpha_channel_ignored(self):
        assert compare_font_color(Font(color="00FF0000"), Font(color="FFFF0000")) is True

    def test_different_rgb_fails(self):
        assert compare_font_color(Font(color="FFFF0000"), Font(color="FF00FF00")) is False

    def test_unset_colors_match(self):
        assert compare_font_color(Font(), Font()) is True

    def test_theme_color_resolves_to_rgb(self):
        # Theme 4 (accent1) is 4472C4 in the default Office theme
        themed = Font(color=Color(theme=4, tint=0.0))
        assert compare_font_color(themed, Font(color="FF4472C4")) is True

    def test_theme_color_with_tint(self):
        # Positive tint lightens toward white: 4472C4 at tint 0.5 -> A1B8E1
        themed = Font(color=Color(theme=4, tint=0.5))
        assert compare_font_color(themed, Font(color="FFA1B8E1")) is True

    def test_theme_color_with_negative_tint(self):
        # Negative tint darkens: 4472C4 at tint -0.5 -> 223962
        # int(0x44*0.5)=0x22, int(0x72*0.5)=0x39, int(0xC4*0.5)=0x62
        themed = Font(color=Color(theme=4, tint=-0.5))
        assert compare_font_color(themed, Font(color="FF223962")) is True


class TestClassifyCells:
    def test_split_regression_vs_modification(self, temp_dir):
        input_path = build_workbook(temp_dir / "in.xlsx", cells={"A1": 1, "A2": 2, "A3": 3})
        golden_path = build_workbook(temp_dir / "gold.xlsx", cells={"A1": 1, "A2": 99, "A3": 3})
        wb_in = openpyxl.load_workbook(input_path, data_only=True)
        wb_gold = openpyxl.load_workbook(golden_path, data_only=True)

        regression, modification = classify_cells_by_modification(
            wb_in, wb_gold, "Model", "A1:A3", False, False, None
        )
        assert regression == ["A1", "A3"]
        assert modification == ["A2"]

    def test_missing_sheet_contributes_no_cells(self, temp_dir):
        input_path = build_workbook(temp_dir / "in.xlsx", sheet_name="Other")
        golden_path = build_workbook(temp_dir / "gold.xlsx")
        wb_in = openpyxl.load_workbook(input_path, data_only=True)
        wb_gold = openpyxl.load_workbook(golden_path, data_only=True)

        regression, modification = classify_cells_by_modification(
            wb_in, wb_gold, "Model", "A1:A3", False, False, None
        )
        assert regression == [] and modification == []


class TestCompareClassifiedCells:
    def test_counts_and_error_messages(self, temp_dir):
        golden_path = build_workbook(temp_dir / "gold.xlsx", cells={"A1": 1, "A2": 99, "A3": 3})
        output_path = build_workbook(temp_dir / "out.xlsx", cells={"A1": 1, "A2": 98, "A3": 4})
        wb_gold = openpyxl.load_workbook(golden_path, data_only=True)
        wb_out = openpyxl.load_workbook(output_path, data_only=True)

        rc, rt, mc, mt, errors = compare_classified_cells(
            wb_gold, wb_out, "Model", ["A1", "A3"], ["A2"], False, False, None
        )
        assert (rc, rt) == (1, 2)  # A1 right, A3 wrong (3 vs 4)
        assert (mc, mt) == (0, 1)  # A2: 98 vs 99 is 1/99 = 1.01% off, outside 1%
        assert any("Regression error at Model!A3" in e for e in errors)
        assert any("Modification error at Model!A2" in e for e in errors)

    def test_missing_output_sheet_scores_all_wrong(self, temp_dir):
        golden_path = build_workbook(temp_dir / "gold.xlsx", cells={"A1": 1})
        output_path = build_workbook(temp_dir / "out.xlsx", sheet_name="Other")
        wb_gold = openpyxl.load_workbook(golden_path, data_only=True)
        wb_out = openpyxl.load_workbook(output_path, data_only=True)

        rc, rt, mc, mt, errors = compare_classified_cells(
            wb_gold, wb_out, "Model", ["A1"], [], False, False, None
        )
        assert (rc, rt, mc, mt) == (0, 1, 0, 0)
        assert errors == ["Model worksheet not found"]

    def test_error_value_falls_back_to_formula_compare(self, temp_dir):
        # Golden and output both show #DIV/0! stored as literal strings; the
        # error triggers the formula-level fallback which compares equal.
        golden_path = build_workbook(temp_dir / "gold.xlsx", cells={"A1": "#DIV/0!"})
        output_path = build_workbook(temp_dir / "out.xlsx", cells={"A1": "#DIV/0!"})
        wb_gold = openpyxl.load_workbook(golden_path, data_only=True)
        wb_out = openpyxl.load_workbook(output_path, data_only=True)
        books = _LazyFormulaWorkbooks(None, golden_path, output_path)

        rc, rt, mc, mt, errors = compare_classified_cells(
            wb_gold, wb_out, "Model", [], ["A1"], False, False, books
        )
        assert (mc, mt) == (1, 1)

    def test_lazy_books_not_loaded_without_errors(self, temp_dir):
        golden_path = build_workbook(temp_dir / "gold.xlsx", cells={"A1": 1})
        output_path = build_workbook(temp_dir / "out.xlsx", cells={"A1": 1})
        wb_gold = openpyxl.load_workbook(golden_path, data_only=True)
        wb_out = openpyxl.load_workbook(output_path, data_only=True)
        books = _LazyFormulaWorkbooks(None, golden_path, output_path)

        compare_classified_cells(wb_gold, wb_out, "Model", ["A1"], [], False, False, books)
        assert books._books == {}  # nothing loaded


class TestCompareWorkbooks:
    def make_files(self, temp_dir, input_cells, golden_cells, output_cells):
        return (
            build_workbook(temp_dir / "in.xlsx", cells=input_cells),
            build_workbook(temp_dir / "gold.xlsx", cells=golden_cells),
            build_workbook(temp_dir / "out.xlsx", cells=output_cells),
        )

    def test_output_equals_golden_passes(self, temp_dir):
        cells = {"A1": 1, "A2": 2, "A3": 3}
        golden = {"A1": 1, "A2": 99, "A3": 3}
        inp, gold, out = self.make_files(temp_dir, cells, golden, dict(golden))
        result = compare_workbooks(inp, gold, out, [("Model", "A1:A3")])
        assert result.passed is True
        assert result.regression_accuracy == 1.0
        assert result.modification_accuracy == 1.0
        assert result.message == ""

    def test_output_equals_input_fails_modification(self, temp_dir):
        cells = {"A1": 1, "A2": 2, "A3": 3}
        golden = {"A1": 1, "A2": 99, "A3": 3}
        inp, gold, out = self.make_files(temp_dir, cells, golden, dict(cells))
        result = compare_workbooks(inp, gold, out, [("Model", "A1:A3")])
        assert result.passed is False
        assert result.regression_accuracy == 1.0
        assert result.modification_accuracy == 0.0
        assert "Modification error at Model!A2" in result.message
        assert "0/2 regression and 1/1 modification cells wrong" in result.message

    def test_regression_snap_at_998(self, temp_dir):
        # 500 A-cells with values; golden adds B1 as the modification cell.
        # Empty B2..B500 classify as regression (None == None), so regression
        # total = 999; output gets A1 wrong -> 998/999 = 0.999 -> snaps to 1.0
        n = 500
        cells = {f"A{i}": i for i in range(1, n + 1)}
        golden = dict(cells)
        golden["B1"] = 42
        output = dict(golden)
        output["A1"] = -1
        inp, gold, out = self.make_files(temp_dir, cells, golden, output)
        result = compare_workbooks(inp, gold, out, [("Model", f"A1:B{n}")])
        assert result.regression_accuracy == 1.0
        assert result.modification_accuracy == 1.0
        assert result.passed is True

    def test_regression_below_snap_fails(self, temp_dir):
        # 3 of 100 regression cells wrong -> 0.97 -> no snap
        cells = {f"A{i}": i for i in range(1, 101)}
        golden = dict(cells)
        output = dict(cells)
        for i in (1, 2, 3):
            output[f"A{i}"] = -i
        inp, gold, out = self.make_files(temp_dir, cells, golden, output)
        result = compare_workbooks(inp, gold, out, [("Model", "A1:A100")])
        assert result.passed is False
        assert result.regression_accuracy == 0.97

    def test_zero_total_modification_group_scores_zero(self, temp_dir):
        # Input already equals golden: no modification cells -> mod ratio 0.0,
        # task can never pass (upstream behavior, ported verbatim)
        cells = {"A1": 1}
        inp, gold, out = self.make_files(temp_dir, cells, dict(cells), dict(cells))
        result = compare_workbooks(inp, gold, out, [("Model", "A1")])
        assert result.passed is False
        assert result.regression_accuracy == 1.0
        assert result.modification_accuracy == 0.0

    def test_multiple_ranges_aggregate(self, temp_dir):
        inp = build_workbook(temp_dir / "in.xlsx", cells={"A1": 1, "B1": 2})
        gold = build_workbook(temp_dir / "gold.xlsx", cells={"A1": 10, "B1": 20})
        out = build_workbook(temp_dir / "out.xlsx", cells={"A1": 10, "B1": 2})
        result = compare_workbooks(inp, gold, out, [("Model", "A1"), ("Model", "B1")])
        assert result.modification_accuracy == 0.5
        assert result.passed is False


def make_v2_task(
    dataset_dir,
    input_cells,
    golden_cells,
    answer_position="'Model'!A1:A3",
    input_name="01_01_input.xlsx",
):
    """Build a v2 dataset dir with one task's files and return the Task."""
    sheets_dir = dataset_dir / "spreadsheet" / "01_proj"
    sheets_dir.mkdir(parents=True, exist_ok=True)
    build_workbook(sheets_dir / input_name, cells=input_cells)
    build_workbook(sheets_dir / "01_golden.xlsx", cells=golden_cells)
    return Task(
        id="01_01",
        instruction="test",
        spreadsheet_path=f"spreadsheet/01_proj/{input_name}",
        answer_position=answer_position,
        golden_response_path="spreadsheet/01_proj/01_golden.xlsx",
    )


class TestEvaluatorV2Dispatch:
    def test_v2_task_graded_with_v2_semantics(self, temp_dir):
        # Output differs from golden within 1% -> v2 passes where v1 would fail
        task = make_v2_task(
            temp_dir,
            input_cells={"A1": 1, "A2": 2, "A3": 3},
            golden_cells={"A1": 1, "A2": 100.0, "A3": 3},
        )
        output = build_workbook(temp_dir / "out.xlsx", cells={"A1": 1, "A2": 100.5, "A3": 3})

        result = Evaluator(temp_dir).evaluate(task, output)
        assert result.passed is True
        assert result.regression_accuracy == 1.0
        assert result.modification_accuracy == 1.0

    def test_v1_task_untouched(self, temp_dir):
        # v1 tasks still route to strict grading and carry no ratios
        task_dir = temp_dir / "spreadsheet" / "13-1"
        task_dir.mkdir(parents=True)
        build_workbook(task_dir / "1_13-1_golden.xlsx", sheet_name="Sheet1", cells={"C1": 100.0})
        output = build_workbook(temp_dir / "out.xlsx", sheet_name="Sheet1", cells={"C1": 100.5})

        result = Evaluator(temp_dir).evaluate(
            Task(
                id="13-1",
                instruction="t",
                spreadsheet_path="spreadsheet/13-1",
                answer_position="C1",
                answer_sheet="Sheet1",
            ),
            output,
        )
        assert result.passed is False  # 100.5 != 100.0 under strict grading
        assert result.regression_accuracy is None
        assert result.modification_accuracy is None

    def test_unreadable_input_scores_zero(self, temp_dir):
        task = make_v2_task(
            temp_dir,
            input_cells={"A1": 1},
            golden_cells={"A1": 2},
        )
        # Corrupt the input file (mirrors the DigiMark malformed-XML files)
        (temp_dir / "spreadsheet" / "01_proj" / "01_01_input.xlsx").write_bytes(b"not a zip")
        output = build_workbook(temp_dir / "out.xlsx", cells={"A1": 2})

        result = Evaluator(temp_dir).evaluate(task, output)
        assert result.passed is False
        assert result.regression_accuracy == 0.0
        assert result.modification_accuracy == 0.0
        assert "Evaluation error" in result.message

    def test_missing_golden_scores_zero_for_v2(self, temp_dir):
        task = make_v2_task(temp_dir, input_cells={"A1": 1}, golden_cells={"A1": 2})
        (temp_dir / "spreadsheet" / "01_proj" / "01_golden.xlsx").unlink()
        output = build_workbook(temp_dir / "out.xlsx", cells={"A1": 2})

        result = Evaluator(temp_dir).evaluate(task, output)
        assert result.passed is False
        assert result.regression_accuracy == 0.0
        assert result.modification_accuracy == 0.0
        assert "Golden file not found" in result.message

    def test_debugging_embedded_uses_formula_mode(self, temp_dir):
        # Dataset dir named *Debugging* + 'Embedded' in spreadsheet_path.
        # Range spans A1:A2 so the regression group is non-empty: A2 is empty
        # everywhere (compare_cell_formula(None, None) -> regression, correct),
        # while A1 is the modification cell. A single-cell range would leave
        # the regression group empty, which scores 0.0 and can never pass.
        dataset_dir = temp_dir / "Debugging"
        dataset_dir.mkdir()
        sheets_dir = dataset_dir / "spreadsheet" / "Embedded_case"
        sheets_dir.mkdir(parents=True)
        build_workbook(sheets_dir / "x_input.xlsx", cells={"A1": "=SUM(B1:B2)"})
        build_workbook(sheets_dir / "x_golden.xlsx", cells={"A1": "=SUM(B1:B3)"})
        output = build_workbook(temp_dir / "out.xlsx", cells={"A1": "=sum($B$1:B3)"})

        task = Task(
            id="e1",
            instruction="t",
            spreadsheet_path="spreadsheet/Embedded_case/x_input.xlsx",
            answer_position="'Model'!A1:A2",
            golden_response_path="spreadsheet/Embedded_case/x_golden.xlsx",
        )
        result = Evaluator(dataset_dir).evaluate(task, output)
        # Formula mode: normalized output formula matches golden -> modification correct
        assert result.passed is True
        assert result.regression_accuracy == 1.0
        assert result.modification_accuracy == 1.0

    def test_debugging_color_uses_font_color_mode(self, temp_dir):
        # Dataset dir named *Debugging* + 'Color' in spreadsheet_path.
        # Input A1 has the same VALUE as golden but no font set, while golden's
        # A1 is red: in font-color mode that makes A1 a modification cell (the
        # font differs), which is what proves the mode is active. A2 is empty
        # everywhere with matching (unset) fonts, so it's the regression cell
        # and keeps the regression group non-empty.
        dataset_dir = temp_dir / "Debugging"
        dataset_dir.mkdir()
        sheets_dir = dataset_dir / "spreadsheet" / "Color_case"
        sheets_dir.mkdir(parents=True)
        build_workbook(sheets_dir / "x_input.xlsx", cells={"A1": 1})
        build_workbook(
            sheets_dir / "x_golden.xlsx",
            cells={"A1": 1},
            fonts={"A1": Font(color="FFFF0000")},
        )
        task = Task(
            id="c1",
            instruction="t",
            spreadsheet_path="spreadsheet/Color_case/x_input.xlsx",
            answer_position="'Model'!A1:A2",
            golden_response_path="spreadsheet/Color_case/x_golden.xlsx",
        )

        wrong_output = build_workbook(
            temp_dir / "out_wrong.xlsx", cells={"A1": 1}, fonts={"A1": Font(color="FF00FF00")}
        )
        result = Evaluator(dataset_dir).evaluate(task, wrong_output)
        # Value alone (1 == 1) would pass; font-color mode catches the wrong font
        assert result.passed is False
        assert result.modification_accuracy == 0.0

        correct_output = build_workbook(
            temp_dir / "out_correct.xlsx", cells={"A1": 1}, fonts={"A1": Font(color="FFFF0000")}
        )
        result2 = Evaluator(dataset_dir).evaluate(task, correct_output)
        assert result2.passed is True
        assert result2.regression_accuracy == 1.0
        assert result2.modification_accuracy == 1.0

    def test_embedded_path_outside_debugging_dataset_stays_value_mode(self, temp_dir):
        # Same file layout as test_debugging_embedded_uses_formula_mode, but the
        # dataset dir is NOT named Debugging -> formula mode must not activate.
        dataset_dir = temp_dir / "Financial_Model"
        dataset_dir.mkdir()
        sheets_dir = dataset_dir / "spreadsheet" / "Embedded_case"
        sheets_dir.mkdir(parents=True)
        build_workbook(sheets_dir / "x_input.xlsx", cells={"A1": "=SUM(B1:B2)"})
        build_workbook(sheets_dir / "x_golden.xlsx", cells={"A1": "=SUM(B1:B3)"})
        output = build_workbook(temp_dir / "out.xlsx", cells={"A1": "=sum($B$1:B3)"})

        task = Task(
            id="e2",
            instruction="t",
            spreadsheet_path="spreadsheet/Embedded_case/x_input.xlsx",
            answer_position="'Model'!A1:A2",
            golden_response_path="spreadsheet/Embedded_case/x_golden.xlsx",
        )
        result = Evaluator(dataset_dir).evaluate(task, output)
        # Value mode (data_only=True): formulas have no cached value, so A1 and
        # A2 both read as None == None and classify as regression; with no
        # modification cells the group scores 0.0 and the task can't pass.
        assert result.passed is False
        assert result.modification_accuracy == 0.0


class TestTaskResultRatios:
    def test_ratios_in_results_dict_when_set(self):
        r = TaskResult(
            task_id="01_01", result="fail", regression_accuracy=0.9987, modification_accuracy=0.5
        )
        d = r.to_results_dict()
        assert d["regression_accuracy"] == 0.9987
        assert d["modification_accuracy"] == 0.5

    def test_ratios_omitted_when_none(self):
        d = TaskResult(task_id="13-1", result="pass").to_results_dict()
        assert "regression_accuracy" not in d
        assert "modification_accuracy" not in d
