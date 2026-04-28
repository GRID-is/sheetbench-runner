"""Tests for workbook evaluation."""

import datetime
from pathlib import Path

import openpyxl
import pytest

from sheetbench_runner.entities import Task
from sheetbench_runner.evaluator import (
    Evaluator,
    _col_name_to_num,
    _col_num_to_name,
    _compare_cell_values,
    _generate_cell_names,
    _parse_sheet_cell_ranges,
    _transform_value,
)


class TestTransformValue:
    """Tests for value transformation."""

    def test_integer_to_float(self):
        assert _transform_value(5) == 5.0

    def test_float_rounds_to_two_decimals(self):
        assert _transform_value(3.14159) == 3.14

    def test_string_number_parses_and_rounds(self):
        assert _transform_value("3.14159") == 3.14

    def test_string_non_number_unchanged(self):
        assert _transform_value("hello") == "hello"

    def test_datetime_converts_to_serial(self):
        # Excel serial for 2024-01-15 is approximately 45306
        dt = datetime.datetime(2024, 1, 15)
        result = _transform_value(dt)
        assert isinstance(result, float)
        assert result == 45306.0

    def test_time_converts_to_string_truncated(self):
        # Original code does str(t)[:-3] which removes last 3 chars
        # With microseconds: '14:30:45.123456'[:-3] = '14:30:45.123'
        t = datetime.time(14, 30, 45, 123456)
        assert _transform_value(t) == "14:30:45.123"

    def test_time_without_microseconds_truncated(self):
        # Without microseconds: '14:30:45'[:-3] = '14:30'
        # (This is arguably a bug in the original, but we maintain compatibility)
        t = datetime.time(14, 30, 45)
        assert _transform_value(t) == "14:30"


class TestCompareCellValues:
    """Tests for cell value comparison."""

    def test_equal_numbers(self):
        assert _compare_cell_values(5, 5) is True

    def test_numbers_within_rounding(self):
        assert _compare_cell_values(3.141, 3.142) is True  # Both round to 3.14

    def test_none_and_empty_string_equal(self):
        assert _compare_cell_values(None, "") is True
        assert _compare_cell_values("", None) is True

    def test_both_none_equal(self):
        assert _compare_cell_values(None, None) is True

    def test_both_empty_equal(self):
        assert _compare_cell_values("", "") is True

    def test_type_mismatch_string_vs_none(self):
        # String "hello" doesn't parse as number, so type mismatch with int
        assert _compare_cell_values("hello", 5) is False

    def test_string_numbers_compared_as_numbers(self):
        # Both "5" and 5 become 5.0 after transform, but type check fails
        # Actually, both become float 5.0, let me check the logic...
        # _transform_value("5") -> 5.0 (float)
        # _transform_value(5) -> 5.0 (float)
        # type(5.0) == type(5.0) -> True
        assert _compare_cell_values("5", 5) is True


class TestColumnConversion:
    """Tests for Excel column name/number conversion."""

    def test_col_name_to_num_single_letter(self):
        assert _col_name_to_num("A") == 1
        assert _col_name_to_num("Z") == 26

    def test_col_name_to_num_double_letter(self):
        assert _col_name_to_num("AA") == 27
        assert _col_name_to_num("AB") == 28
        assert _col_name_to_num("AZ") == 52

    def test_col_num_to_name_single_letter(self):
        assert _col_num_to_name(1) == "A"
        assert _col_num_to_name(26) == "Z"

    def test_col_num_to_name_double_letter(self):
        assert _col_num_to_name(27) == "AA"
        assert _col_num_to_name(28) == "AB"
        assert _col_num_to_name(52) == "AZ"

    def test_round_trip(self):
        for name in ["A", "Z", "AA", "AZ", "BA", "ZZ"]:
            assert _col_num_to_name(_col_name_to_num(name)) == name


class TestGenerateCellNames:
    """Tests for cell range expansion."""

    def test_single_cell(self):
        assert _generate_cell_names("A1") == ["A1"]

    def test_simple_range(self):
        names = _generate_cell_names("A1:A3")
        assert names == ["A1", "A2", "A3"]

    def test_multi_column_range(self):
        names = _generate_cell_names("A1:B2")
        assert set(names) == {"A1", "A2", "B1", "B2"}


class TestParseSheetCellRanges:
    """Tests for parsing answer_position into (sheet_name, cell_range) tuples."""

    def test_simple_sheet_name_with_exclamation(self):
        # Arrange & Act
        result = _parse_sheet_cell_ranges("Sheet1!A1:B10")

        # Assert
        assert result == [("Sheet1", "A1:B10")]

    def test_quoted_sheet_name(self):
        # Arrange & Act
        result = _parse_sheet_cell_ranges("'Sheet1'!A1:B10")

        # Assert
        assert result == [("Sheet1", "A1:B10")]

    def test_multiple_ranges_simple_sheets(self):
        # Arrange & Act
        result = _parse_sheet_cell_ranges("Sheet1!A1,Sheet2!B2")

        # Assert
        assert result == [("Sheet1", "A1"), ("Sheet2", "B2")]

    def test_fallback_to_answer_sheet(self):
        # Arrange & Act
        result = _parse_sheet_cell_ranges("A1:B10", answer_sheet="MySheet")

        # Assert
        assert result == [("MySheet", "A1:B10")]

    def test_fallback_to_default_sheet(self):
        # Arrange & Act
        result = _parse_sheet_cell_ranges("A1:B10", fallback_sheet_name="DefaultSheet")

        # Assert
        assert result == [("DefaultSheet", "A1:B10")]

    # ============================================================
    # Tests for sheet names containing commas
    # These verify that quoted sheet names with commas are parsed correctly.
    # ============================================================

    def test_sheet_name_with_single_comma(self):
        """Sheet names with commas are correctly parsed when quoted."""
        # Arrange & Act
        result = _parse_sheet_cell_ranges("'a, b'!A1:B10")

        # Assert
        assert result == [("a, b", "A1:B10")]

    def test_sheet_name_with_multiple_commas(self):
        """Sheet names with multiple commas are correctly parsed when quoted."""
        # Arrange & Act
        result = _parse_sheet_cell_ranges("'a, b, c'!A1")

        # Assert
        assert result == [("a, b, c", "A1")]

    def test_multiple_ranges_with_comma_in_sheet_name(self):
        """Multiple ranges where one sheet name contains a comma."""
        # Arrange & Act
        result = _parse_sheet_cell_ranges("'Sales, Q1'!A1,Sheet2!B2")

        # Assert
        assert result == [("Sales, Q1", "A1"), ("Sheet2", "B2")]

    def test_answer_sheet_with_comma(self):
        """answer_sheet field with comma in name is handled correctly."""
        # Arrange & Act
        result = _parse_sheet_cell_ranges("A1:B10", answer_sheet="'Revenue, 2024'")

        # Assert
        assert result == [("Revenue, 2024", "A1:B10")]

    # ============================================================
    # Tests for malformed quotes (missing opening quote)
    # These occur in real datasets and need repair.
    # ============================================================

    def test_missing_opening_quote_first_range(self):
        """First range missing opening quote, second is correct."""
        # Arrange & Act
        result = _parse_sheet_cell_ranges("Sheet3'!A:G,'Sheet4'!A:G")

        # Assert
        assert result == [("Sheet3", "A:G"), ("Sheet4", "A:G")]

    def test_missing_opening_quotes_multiple_ranges(self):
        """Multiple ranges with missing opening quotes, one unquoted."""
        # Arrange & Act
        result = _parse_sheet_cell_ranges("SOME SHEET'!A2:C100,'OTHER SHEET'!AB44,EXTRA!C5:D5")

        # Assert
        assert result == [
            ("SOME SHEET", "A2:C100"),
            ("OTHER SHEET", "AB44"),
            ("EXTRA", "C5:D5"),
        ]

    def test_quotes_around_entire_reference(self):
        """Quotes wrapping entire reference instead of just sheet name."""
        # Arrange & Act
        result = _parse_sheet_cell_ranges("'9045!A1:F9','34354!A1:F1'")

        # Assert
        assert result == [("9045", "A1:F9"), ("34354", "A1:F1")]

    def test_external_workbook_reference(self):
        """External workbook reference with quoted sheet name."""
        # Arrange & Act
        result = _parse_sheet_cell_ranges("[workbook2.xlsx]'sheet 2'!B4")

        # Assert
        assert result == [("[workbook2.xlsx]sheet 2", "B4")]

    # ============================================================
    # Error cases
    # ============================================================

    def test_no_sheet_name_raises_error(self):
        """Raises ValueError when no sheet name is available."""
        # Arrange & Act & Assert
        with pytest.raises(ValueError, match="No sheet name available"):
            _parse_sheet_cell_ranges("A1:B10")


@pytest.fixture
def evaluator_setup(temp_dir: Path):
    """Set up a dataset directory with golden file for testing."""
    dataset_dir = temp_dir / "dataset"
    spreadsheet_dir = dataset_dir / "spreadsheet" / "test-1"
    spreadsheet_dir.mkdir(parents=True)

    # Create a golden workbook
    golden_wb = openpyxl.Workbook()
    golden_ws = golden_wb.active
    golden_ws.title = "Sheet1"
    golden_ws["A1"] = 10
    golden_ws["A2"] = 20
    golden_ws["A3"] = 30
    golden_wb.save(spreadsheet_dir / "1_test-1_golden.xlsx")

    return dataset_dir, spreadsheet_dir


def test_evaluator_pass(evaluator_setup, temp_dir: Path):
    """Test evaluation passes when values match."""
    # Arrange
    dataset_dir, spreadsheet_dir = evaluator_setup

    task = Task(
        id="test-1",
        instruction="Test",
        spreadsheet_path="spreadsheet/test-1",
        instruction_type="Cell-Level Manipulation",
        answer_position="A1:A3",
    )

    # Create matching output
    output_wb = openpyxl.Workbook()
    output_ws = output_wb.active
    output_ws.title = "Sheet1"
    output_ws["A1"] = 10
    output_ws["A2"] = 20
    output_ws["A3"] = 30
    output_path = temp_dir / "output.xlsx"
    output_wb.save(output_path)

    evaluator = Evaluator(dataset_dir)

    # Act
    result = evaluator.evaluate(task, output_path)

    # Assert
    assert result.passed is True
    assert result.message == ""


def test_evaluator_fail_value_mismatch(evaluator_setup, temp_dir: Path):
    """Test evaluation fails when values don't match."""
    # Arrange
    dataset_dir, spreadsheet_dir = evaluator_setup

    task = Task(
        id="test-1",
        instruction="Test",
        spreadsheet_path="spreadsheet/test-1",
        instruction_type="Cell-Level Manipulation",
        answer_position="A1:A3",
    )

    # Create non-matching output
    output_wb = openpyxl.Workbook()
    output_ws = output_wb.active
    output_ws.title = "Sheet1"
    output_ws["A1"] = 10
    output_ws["A2"] = 999  # Wrong!
    output_ws["A3"] = 30
    output_path = temp_dir / "output.xlsx"
    output_wb.save(output_path)

    evaluator = Evaluator(dataset_dir)

    # Act
    result = evaluator.evaluate(task, output_path)

    # Assert
    assert result.passed is False
    assert "A2" in result.message
    assert "999" in result.message


def test_evaluator_missing_output_file(evaluator_setup, temp_dir: Path):
    """Test evaluation fails when output file doesn't exist."""
    # Arrange
    dataset_dir, _ = evaluator_setup

    task = Task(
        id="test-1",
        instruction="Test",
        spreadsheet_path="spreadsheet/test-1",
        instruction_type="Cell-Level Manipulation",
        answer_position="A1:A3",
    )

    evaluator = Evaluator(dataset_dir)

    # Act
    result = evaluator.evaluate(task, temp_dir / "nonexistent.xlsx")

    # Assert
    assert result.passed is False
    assert "not found" in result.message.lower()
